import streamlit as st
import requests
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import time
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

API_URL = "http://127.0.0.1:8002"

def esperar_api():
    for _ in range(15):
        try:
            r = requests.get(f"{API_URL}/api/v1/salud", timeout=2)
            if r.status_code == 200:
                return True
        except:
            pass
        time.sleep(1)
    return False

if 'api_conectada' not in st.session_state:
    st.session_state.api_conectada = esperar_api()

st.set_page_config(page_title="MarkeTTalento", page_icon="📦", layout="wide")

def api_get(endpoint):
    try:
        r = requests.get(f"{API_URL}{endpoint}", timeout=5)
        return r.json() if r.status_code == 200 else []
    except:
        return []

def api_post(endpoint, data):
    try:
        r = requests.post(f"{API_URL}{endpoint}", json=data, timeout=5)
        if r.status_code in [200, 201]:
            return r.json()
        else:
            print(f"API Error: {r.status_code} - {r.text}")
            return {}
    except Exception as e:
        print(f"API Exception: {e}")
        return {}

def api_put(endpoint, data):
    try:
        r = requests.put(f"{API_URL}{endpoint}", json=data, timeout=5)
        if r.status_code in [200, 201]:
            return r.json()
        else:
            error_msg = f"API Error: {r.status_code} - {r.text}"
            print(error_msg)
            return {"error": error_msg}
    except Exception as e:
        error_msg = f"API Exception: {e}"
        print(error_msg)
        return {"error": error_msg}

def conectar_api():
    try:
        r = requests.get(f"{API_URL}/api/v1/salud", timeout=3)
        return r.status_code == 200
    except:
        return False

if not conectar_api():
    st.error("❌ No conectado a la API. Inicia 'python iniciar.py'")
    st.stop()

def to_excel(df):
    if df is None or df.empty:
        return b""
    
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    try:
        ws.title = "Datos"
    except:
        pass
    
    header_fill = PatternFill(start_color="00f0ff", end_color="00f0ff", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    wb.save(output)
    return output.getvalue()

st.markdown("""
<style>
    /* Importar fuentes */
    @import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700&family=Rajdhani:wght@300;400;500;600;700&display=swap');
    @import url('https://cdn.jsdelivr.net/npm/@fontsource/cascadia-code@4.2.1/index.min.css');
    @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=swap');
    
    /* Tema oscuro futurista */
    :root {
        --bg-primary: #0a0e17;
        --bg-secondary: #111827;
        --bg-card: #1a2332;
        --accent-cyan: #00f0ff;
        --accent-purple: #8b5cf6;
        --accent-pink: #ec4899;
        --accent-green: #10b981;
        --accent-red: #ef4444;
        --accent-orange: #f59e0b;
        --text-primary: #f1f5f9;
        --text-secondary: #94a3b8;
        --border-glow: rgba(0, 240, 255, 0.3);
    }
    
    /* Fondo principal */
    .stApp {
        background: linear-gradient(135deg, var(--bg-primary) 0%, var(--bg-secondary) 100%);
        color: var(--text-primary);
    }
    
    /* Tarjetas de métricas */
    .metric-card {
        background: linear-gradient(135deg, var(--bg-card) 0%, rgba(26, 35, 50, 0.8) 100%);
        border: 1px solid rgba(0, 240, 255, 0.2);
        border-radius: 16px;
        padding: 20px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.05);
        transition: all 0.3s ease;
    }
    
    .metric-card:hover {
        border-color: var(--accent-cyan);
        box-shadow: 0 4px 30px rgba(0, 240, 255, 0.15);
        transform: translateY(-2px);
    }
    
    /* Títulos */
    h1, h2, h3 {
        font-family: 'Cascadia Code', 'Orbitron', monospace;
        color: var(--accent-cyan);
        text-shadow: 0 0 20px rgba(0, 240, 255, 0.3);
    }
    
    /* Fuente global - excluir elementos de iconos del sidebar */
    html, body, .stApp, [class*="st"] {
        font-family: 'Cascadia Code', 'Segoe UI', monospace !important;
    }
    
    /* Excluir específicamente los elementos de iconos de Streamlit */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarCollapseButton"] *,
    .material-symbols-rounded,
    .material-symbols-outlined,
    .st-emotion-cache-1cypcdb,  /* Posible clase del botón */
    button[kind="header"] {
        font-family: 'Material Symbols Rounded', sans-serif !important;
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, var(--bg-card) 0%, var(--bg-primary) 100%);
        border-right: 1px solid rgba(0, 240, 255, 0.1);
    }
    
    /* Botón colapsar sidebar - usar Material Symbols Rounded para los iconos */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarCollapseButton"] span,
    [data-testid="stSidebarCollapseButton"] button {
        font-family: 'Material Symbols Rounded', sans-serif !important;
        font-size: 24px !important;
    }
    
    /* Asegurar que el icono se muestre correctamente */
    .material-symbols-rounded,
    .material-symbols-outlined {
        font-family: 'Material Symbols Rounded' !important;
    }
    

    
    /* Botones principales */
    .stButton > button:first-child {
        background: linear-gradient(135deg, var(--accent-cyan) 0%, var(--accent-purple) 100%);
        color: var(--bg-primary);
        font-weight: 600;
        border: none;
        border-radius: 10px;
        padding: 10px 25px;
        box-shadow: 0 4px 15px rgba(0, 240, 255, 0.3);
        transition: all 0.3s ease;
    }
    
    .stButton > button:first-child:hover {
        box-shadow: 0 6px 25px rgba(0, 240, 255, 0.5);
        transform: translateY(-2px);
    }
    
    /* Tabs */
    .stTabs [data-selected] {
        background: linear-gradient(135deg, var(--accent-cyan), var(--accent-purple));
        color: var(--bg-primary) !important;
        border-radius: 10px 10px 0 0;
    }
    
    /* Inputs */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > input {
        background: var(--bg-card);
        border: 1px solid rgba(0, 240, 255, 0.2);
        color: var(--text-primary);
        border-radius: 10px;
    }
    
    /* File Uploader - estilo limpio */
    [data-testid="stFileUploader"] {
        background: rgba(30, 41, 59, 0.6);
        border: 1px solid rgba(0, 240, 255, 0.2);
        border-radius: 10px;
        padding: 10px;
    }
    
    [data-testid="stFileUploader"] > section {
        border: 2px dashed rgba(0, 240, 255, 0.3) !important;
        background: rgba(0, 0, 0, 0.2) !important;
        border-radius: 8px !important;
    }
    
    [data-testid="stFileUploader"] button {
        background: linear-gradient(135deg, var(--accent-cyan), var(--accent-purple)) !important;
        color: var(--bg-primary) !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        padding: 6px 16px !important;
    }
    
    [data-testid="stFileUploader"] button:hover {
        box-shadow: 0 4px 15px rgba(0, 240, 255, 0.4) !important;
    }
    
    [data-testid="stFileUploader"] span {
        color: var(--text-secondary) !important;
        font-size: 0.8rem !important;
    }
    
    /* Dataframes */
    [data-testid="stDataFrame"] {
        background: var(--bg-card);
        border-radius: 12px;
    }
    
    /* Expander */
    .streamlit-expander {
        background: var(--bg-card);
        border: 1px solid rgba(0, 240, 255, 0.1);
        border-radius: 12px;
    }
    
    /* Success/Error alerts */
    .stSuccess, .stError, .stWarning, .stInfo {
        border-radius: 12px;
    }
    
    /* Spinner */
    .stSpinner > div {
        border-color: var(--accent-cyan) !important;
    }
    
    /* Hide default streamlit elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Scrollbar personalizado */
    ::-webkit-scrollbar {
        width: 8px;
    }
    ::-webkit-scrollbar-track {
        background: var(--bg-primary);
    }
    ::-webkit-scrollbar-thumb {
        background: var(--accent-cyan);
        border-radius: 4px;
    }
    
    /* Status indicators */
    .status-dot {
        display: inline-block;
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 8px;
        animation: pulse 2s infinite;
    }
    
    @keyframes pulse {
        0% { opacity: 1; }
        50% { opacity: 0.5; }
        100% { opacity: 1; }
    }
    
    /* Card glow effect */
    .glow-card {
        position: relative;
    }
    
    .glow-card::before {
        content: '';
        position: absolute;
        top: -1px;
        left: -1px;
        right: -1px;
        bottom: -1px;
        background: linear-gradient(45deg, var(--accent-cyan), var(--accent-purple), var(--accent-pink));
        border-radius: 17px;
        z-index: -1;
        opacity: 0;
        transition: opacity 0.3s ease;
    }
    
    .glow-card:hover::before {
        opacity: 1;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<style>.hdr-c{background:linear-gradient(135deg,rgba(0,240,255,0.15) 0%,rgba(139,92,246,0.15) 50%,rgba(236,72,153,0.15) 100%);padding:30px 40px;border-radius:20px;border:1px solid rgba(0,240,255,0.3);box-shadow:0 0 40px rgba(0,240,255,0.2),inset 0 0 30px rgba(0,240,255,0.05);margin-bottom:25px;position:relative;overflow:hidden;}.hdr-t,.hdr-b{position:absolute;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent,#00f0ff,transparent);box-shadow:0 0 20px rgba(0,240,255,0.8);}.hdr-t{top:0;}.hdr-b{bottom:0;background:linear-gradient(90deg,transparent,#8b5cf6,transparent);box-shadow:0 0 20px rgba(139,92,246,0.8);}.hdr-x{text-align:center;position:relative;z-index:1;}.hdr-i{font-size:3.5rem;filter:drop-shadow(0 0 15px rgba(0,240,255,0.6));}.hdr-h{font-size:3rem;margin:15px 0 10px 0;font-weight:700;letter-spacing:3px;background:linear-gradient(90deg,#00f0ff,#8b5cf6,#ec4899);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}.hdr-p{color:#94a3b8;font-size:1.2rem;margin:0;font-weight:500;letter-spacing:2px;}.hdr-v{color:#00f0ff;text-shadow:0 0 10px rgba(0,240,255,0.5);}.hdr-bad{margin-top:20px;display:flex;justify-content:center;gap:15px;flex-wrap:wrap;}.hdr-b1,.hdr-b2,.hdr-b3{padding:6px 16px;border-radius:20px;font-size:0.85rem;font-weight:600;}.hdr-b1{background:rgba(0,240,255,0.15);border:1px solid rgba(0,240,255,0.4);color:#00f0ff;}.hdr-b2{background:rgba(139,92,246,0.15);border:1px solid rgba(139,92,246,0.4);color:#8b5cf6;}.hdr-b3{background:rgba(16,185,129,0.15);border:1px solid rgba(16,185,129,0.4);color:#10b981;}</style><div class="hdr-c"><div class="hdr-t"></div><div class="hdr-b"></div><div class="hdr-x"><span class="hdr-i">📦</span><h1 class="hdr-h">MarkeTTalento</h1><p class="hdr-p">Sistema de Inventario Inteligente con <span class="hdr-v">Visión Artificial</span></p><div class="hdr-bad"><span class="hdr-b1">🤖 ML & IA</span><span class="hdr-b2">📊 Analytics</span><span class="hdr-b3">🔮 Predicciones</span></div></div></div>', unsafe_allow_html=True)

st.markdown("---")

# Sidebar
with st.sidebar:
    st.markdown("<h2 style='text-align: center; font-family: \"Cascadia Code\", \"Orbitron\", monospace;'>⚡ Menú</h2>", unsafe_allow_html=True)
    
    st.markdown("""
    <style>
        .menu-card {
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 10px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
            border: 1px solid rgba(255,255,255,0.1);
        }
        .menu-card:hover {
            transform: scale(1.02);
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
        }
        .menu-card a {
            text-decoration: none;
            color: inherit;
        }
        .menu-icon { font-size: 2rem; margin-bottom: 5px; }
        .menu-label { font-weight: 600; font-size: 0.9rem; }
        
        .card-dashboard { background: linear-gradient(135deg, rgba(0, 240, 255, 0.2), rgba(0, 240, 255, 0.05)); border-color: rgba(0, 240, 255, 0.3); }
        .card-productos { background: linear-gradient(135deg, rgba(139, 92, 246, 0.2), rgba(139, 92, 246, 0.05)); border-color: rgba(139, 92, 246, 0.3); }
        .card-inventario { background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(16, 185, 129, 0.05)); border-color: rgba(16, 185, 129, 0.3); }
        .card-ventas { background: linear-gradient(135deg, rgba(245, 158, 11, 0.2), rgba(245, 158, 11, 0.05)); border-color: rgba(245, 158, 11, 0.3); }
        .card-predicciones { background: linear-gradient(135deg, rgba(236, 72, 153, 0.2), rgba(236, 72, 153, 0.05)); border-color: rgba(236, 72, 153, 0.3); }
        .card-vision { background: linear-gradient(135deg, rgba(59, 130, 246, 0.2), rgba(59, 130, 246, 0.05)); border-color: rgba(59, 130, 246, 0.3); }
        .card-link { background: linear-gradient(135deg, rgba(100, 100, 100, 0.2), rgba(100, 100, 100, 0.05)); border-color: rgba(100, 100, 100, 0.3); }
        .card-estado { background: linear-gradient(135deg, rgba(0, 240, 255, 0.1), rgba(0, 0, 0, 0.1)); border-color: rgba(0, 240, 255, 0.2); }
        
        .txt-cyan { color: #00f0ff; }
        .txt-purple { color: #8b5cf6; }
        .txt-green { color: #10b981; }
        .txt-orange { color: #f59e0b; }
        .txt-pink { color: #ec4899; }
.txt-blue { color: #3b82f6; }
        .txt-white { color: #f1f5f9; }
    </style>
    """, unsafe_allow_html=True)
    
    def card(icon, label, color_class, txt_class, selected, key):
        active = "border-width: 2px;" if selected else ""
        st.markdown(f"""
        <div class="menu-card {color_class}" style="{active}">
            <div class="menu-icon">{icon}</div>
            <div class="menu-label {txt_class}">{label}</div>
        </div>
        """, unsafe_allow_html=True)
        return st.button(label, key=key, use_container_width=True)
    
    st.markdown("---")
    
    if 'menu_activo' not in st.session_state:
        st.session_state['menu_activo'] = "🏠 Dashboard"
    
    btn_dashboard = st.button("🏠 Dashboard", use_container_width=True, key="btn_dash")
    btn_productos = st.button("📦 Productos", use_container_width=True, key="btn_prod")
    btn_inventario = st.button("📊 Inventario", use_container_width=True, key="btn_inv")
    btn_ventas = st.button("💰 Ventas", use_container_width=True, key="btn_vent")
    btn_predicciones = st.button("🔮 Predicciones", use_container_width=True, key="btn_pred")
    btn_vision = st.button("📸 Visión AI", use_container_width=True, key="btn_vis")
    btn_barcode = st.button("🔍 Barcode", use_container_width=True, key="btn_bar")
    
    if btn_dashboard:
        st.session_state['menu_activo'] = "🏠 Dashboard"
    elif btn_productos:
        st.session_state['menu_activo'] = "📦 Productos"
    elif btn_inventario:
        st.session_state['menu_activo'] = "📊 Inventario"
    elif btn_ventas:
        st.session_state['menu_activo'] = "💰 Ventas"
    elif btn_predicciones:
        st.session_state['menu_activo'] = "🔮 Predicciones"
    elif btn_vision:
        st.session_state['menu_activo'] = "📸 Visión AI"
    elif btn_barcode:
        st.session_state['menu_activo'] = "🔍 Barcode"
    
    menu = st.session_state['menu_activo']
    
    st.markdown("---")
    
    st.markdown("""
    <div style="padding: 15px; background: linear-gradient(135deg, rgba(16, 185, 129, 0.15), rgba(0, 0, 0, 0.2)); border-radius: 12px; border: 1px solid rgba(16, 185, 129, 0.3); margin-bottom: 10px;">
        <div style="display: flex; align-items: center; margin-bottom: 10px;">
            <span style="width: 10px; height: 10px; background: #10b981; border-radius: 50%; margin-right: 8px; animation: pulse 2s infinite;"></span>
            <span style="color: #10b981; font-weight: 600;">API Online</span>
            <span style="color: #64748b; margin-left: auto; font-size: 0.85rem;">Puerto: 8002</span>
        </div>
        <div id="reloj" style="color: #f1f5f9; font-size: 1.1rem; font-weight: 600; text-align: center;"></div>
        <div id="fecha" style="color: #94a3b8; font-size: 0.85rem; text-align: center;"></div>
    </div>
    <script>
        function actualizarReloj() {
            const now = new Date();
            const dias = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado'];
            const meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'];
            const dia = dias[now.getDay()];
            const fecha = now.getDate();
            const mes = meses[now.getMonth()];
            const anio = now.getFullYear();
            const hora = String(now.getHours()).padStart(2, '0');
            const min = String(now.getMinutes()).padStart(2, '0');
            const seg = String(now.getSeconds()).padStart(2, '0');
            document.getElementById('reloj').textContent = hora + ':' + min + ':' + seg;
            document.getElementById('fecha').textContent = dia + ' ' + fecha + ' ' + mes + ' ' + anio;
        }
        actualizarReloj();
        setInterval(actualizarReloj, 1000);
    </script>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <a href="http://localhost:8002/docs" target="_blank" style="display: block; padding: 10px; background: linear-gradient(135deg, rgba(139, 92, 246, 0.1), rgba(0, 0, 0, 0.1)); border-radius: 10px; border: 1px solid rgba(139, 92, 246, 0.2); text-decoration: none; text-align: center;">
        <span style="color: #8b5cf6;">📚 API Docs</span>
    </a>
    """, unsafe_allow_html=True)


# Dashboard principal
if menu == "🏠 Dashboard":
    st.markdown("<h2>🏠 Dashboard</h2>", unsafe_allow_html=True)

    if 'filtro_agotados' not in st.session_state:
        st.session_state.filtro_agotados = True
    if 'filtro_criticos' not in st.session_state:
        st.session_state.filtro_criticos = True
    if 'filtro_bajos' not in st.session_state:
        st.session_state.filtro_bajos = True
    if 'filtro_saludables' not in st.session_state:
        st.session_state.filtro_saludables = True

    # Calcular estados basados en porcentajes del STOCK MÁXIMO (antes de las métricas)
    inventarios_calc = api_get("/api/v1/inventario")
    productos_calc = api_get("/api/v1/productos")
    
    agotados = 0
    criticos = 0
    bajos = 0
    saludables = 0
    
    for inv in inventarios_calc:
        prod = next((p for p in productos_calc if p.get("id") == inv.get("producto_id")), None)
        if prod:
            stock = inv.get("cantidad", 0)
            max_s = prod.get("stock_maximo", 100) or 100
            
            if stock <= 0:
                agotados += 1
            elif max_s > 0:
                pct = (stock / max_s) * 100
                if pct <= 25:
                    criticos += 1
                elif pct <= 59:
                    bajos += 1
                else:
                    saludables += 1
            else:
                saludables += 1
    
    # Guardar en session_state para uso en otras partes
    st.session_state.dashboard_agotados = agotados
    st.session_state.dashboard_criticos = criticos
    st.session_state.dashboard_bajos = bajos
    st.session_state.dashboard_saludables = saludables
    
    # Obtener resumen del API (para métricas que no dependen de estados)
    resumen = api_get("/api/v1/inventario/resumen")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("""
        <div class="metric-card glow-card" style="padding: 20px; text-align: center; cursor: pointer;" onclick="window.location.href='#productos'">
            <div style="color: #94a3b8; font-size: 0.9rem;">PRODUCTOS</div>
            <div style="font-size: 2.5rem; font-weight: 700; color: #00f0ff; text-shadow: 0 0 20px rgba(0, 240, 255, 0.5);">
                """ + str(resumen.get('total_productos', 0)) + """
            </div>
            <div style="color: #94a3b8; font-size: 1rem; font-weight: 500;">en catálogo</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="metric-card glow-card" style="padding: 20px; text-align: center; cursor: pointer;">
            <div style="color: #94a3b8; font-size: 0.9rem;">STOCK TOTAL</div>
            <div style="font-size: 2.5rem; font-weight: 700; color: #8b5cf6; text-shadow: 0 0 20px rgba(139, 92, 246, 0.5);">
                """ + str(resumen.get('total_unidades', 0)) + """
            </div>
            <div style="color: #94a3b8; font-size: 1rem; font-weight: 500;">unidades</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        color = "#ef4444" if criticos > 0 else "#10b981"
        st.markdown(f"""
        <div class="metric-card glow-card" style="padding: 20px; text-align: center; cursor: pointer;">
            <div style="color: #94a3b8; font-size: 0.9rem;">CRÍTICOS</div>
            <div style="font-size: 2.5rem; font-weight: 700; color: {color}; text-shadow: 0 0 20px {color}40;">
                {criticos}
            </div>
            <div style="color: #94a3b8; font-size: 1rem; font-weight: 500;">requieren atención</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        valor = resumen.get("valor_total", 0)
        st.markdown(f"""
        <div class="metric-card glow-card" style="padding: 20px; text-align: center; cursor: pointer;">
            <div style="color: #94a3b8; font-size: 0.9rem;">VALOR TOTAL</div>
            <div style="font-size: 2.5rem; font-weight: 700; color: #10b981; text-shadow: 0 0 20px rgba(16, 185, 129, 0.5);">
                €{valor:.0f}
            </div>
            <div style="color: #94a3b8; font-size: 1rem; font-weight: 500;">en inventario</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Gráficos
    col_chart1, col_chart2 = st.columns([2, 1])
    
    with col_chart1:
        st.markdown("""
        <div class="metric-card">
            <h3 style="margin-bottom: 15px;">📊 Estado del Inventario</h3>
        </div>
        """, unsafe_allow_html=True)
        
        datos = [
            {"Estado": "Agotados", "Cantidad": agotados, "Color": "#6b7280"},
            {"Estado": "Críticos", "Cantidad": criticos, "Color": "#ef4444"},
            {"Estado": "Bajos", "Cantidad": bajos, "Color": "#f59e0b"},
            {"Estado": "Saludables", "Cantidad": saludables, "Color": "#10b981"},
        ]
        
        if 'filtro_agotados' not in st.session_state:
            st.session_state.filtro_agotados = True
        if 'filtro_criticos' not in st.session_state:
            st.session_state.filtro_criticos = True
        if 'filtro_bajos' not in st.session_state:
            st.session_state.filtro_bajos = True
        if 'filtro_saludables' not in st.session_state:
            st.session_state.filtro_saludables = True
        
        col_f1, col_f2, col_f3, col_f4 = st.columns([1, 1, 1, 1])
        with col_f1:
            filtro_agotados = st.checkbox("Agotados", value=st.session_state.filtro_agotados, key="f_check_agotados")
        with col_f2:
            filtro_criticos = st.checkbox("Críticos", value=st.session_state.filtro_criticos, key="f_check_criticos")
        with col_f3:
            filtro_bajos = st.checkbox("Bajos", value=st.session_state.filtro_bajos, key="f_check_bajos")
        with col_f4:
            filtro_saludables = st.checkbox("Saludables", value=st.session_state.filtro_saludables, key="f_check_saludables")
        
        st.session_state.filtro_agotados = filtro_agotados
        st.session_state.filtro_criticos = filtro_criticos
        st.session_state.filtro_bajos = filtro_bajos
        st.session_state.filtro_saludables = filtro_saludables
        
        datos_filtrados = [d for d in datos if (
            (d["Estado"] == "Agotados" and filtro_agotados) or
            (d["Estado"] == "Críticos" and filtro_criticos) or
            (d["Estado"] == "Bajos" and filtro_bajos) or
            (d["Estado"] == "Saludables" and filtro_saludables)
        )]
        
        if datos_filtrados:
            fig = go.Figure()
            
            fig.add_trace(go.Pie(
                labels=[d["Estado"] for d in datos_filtrados],
                values=[d["Cantidad"] for d in datos_filtrados],
                marker=dict(colors=[d["Color"] for d in datos_filtrados]),
                hole=0.6,
                textinfo='label+percent',
                textposition='outside',
                textfont=dict(size=16, color='white', family='Cascadia Code', weight=600),
                hovertemplate="<b>%{label}</b><br>Cantidad: %{value}<extra></extra>"
            ))
            
            fig.update_layout(
                showlegend=True,
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=-0.25,
                    xanchor="center",
                    x=0.5,
                    font=dict(color="white", size=12)
                ),
                height=400,
                paper_bgcolor="#0a0e17",
                plot_bgcolor="#0a0e17",
                font=dict(color="white"),
                margin=dict(l=20, r=20, t=30, b=80),
                annotations=[dict(
                    text='<b>TOTAL</b>',
                    x=0.5, y=0.5,
                    font_size=14,
                    font_color='white',
                    showarrow=False
                )]
            )
            
            st.plotly_chart(fig, width='stretch')
    
    # Gráfico de barras futurista - stock por producto
    st.markdown("")
    st.markdown("""
    <div style="background: linear-gradient(135deg, rgba(0, 240, 255, 0.1), rgba(139, 92, 246, 0.1)); 
                padding: 20px; border-radius: 15px; border: 1px solid rgba(0, 240, 255, 0.3);
                box-shadow: 0 0 30px rgba(0, 240, 255, 0.1);">
        <h3 style="margin: 0; color: #00f0ff; text-shadow: 0 0 10px rgba(0, 240, 255, 0.5);">
            📦 Stock por Producto
        </h3>
        <p style="margin: 5px 0 0 0; color: #94a3b8; font-size: 0.85rem;">
            Visualización en tiempo real del inventario
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Filtros para Stock por Producto - Sincronizados con el primer gráfico
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    col_filtro1, col_filtro2, col_filtro3, col_filtro4 = st.columns([1, 1, 1, 1])
    with col_filtro1:
        filtro_stock_agotados = st.checkbox("⚫ Agotados", value=st.session_state.filtro_agotados)
        st.session_state.filtro_agotados = filtro_stock_agotados
    with col_filtro2:
        filtro_stock_criticos = st.checkbox("🔴 Críticos", value=st.session_state.filtro_criticos)
        st.session_state.filtro_criticos = filtro_stock_criticos
    with col_filtro3:
        filtro_stock_bajos = st.checkbox("🟡 Bajos", value=st.session_state.filtro_bajos)
        st.session_state.filtro_bajos = filtro_stock_bajos
    with col_filtro4:
        filtro_stock_saludables = st.checkbox("🟢 Saludables", value=st.session_state.filtro_saludables)
        st.session_state.filtro_saludables = filtro_stock_saludables
    
    inventarios = api_get("/api/v1/inventario")
    productos = api_get("/api/v1/productos")
    
    def get_estado_producto(prod_id, inv_list):
        for inv in inv_list:
            if inv.get("producto_id") == prod_id:
                stock = inv.get("cantidad", 0)
                prod = next((p for p in productos if p.get("id") == prod_id), None)
                max_s = prod.get("stock_maximo", 100) or 100 if prod else 100
                
                if stock <= 0:
                    return "Agotados"
                elif max_s > 0:
                    pct = (stock / max_s) * 100
                    if pct <= 25:
                        return "Críticos"
                    elif pct <= 59:
                        return "Bajos"
                    else:
                        return "Saludables"
                else:
                    return "Saludables"
        return "Agotados"
    
    datos_stock = []
    for inv in inventarios:
        prod = next((p for p in productos if p.get("id") == inv.get("producto_id")), None)
        if prod:
            estado = get_estado_producto(prod.get("id"), inventarios)
            datos_stock.append({
                "Producto": prod.get("nombre", "Producto"),
                "Stock": inv.get("cantidad", 0),
                "Estado": estado
            })
    
    datos_stock_filtrados = []
    for d in datos_stock:
        estado = d["Estado"]
        if estado == "Agotados" and filtro_stock_agotados:
            datos_stock_filtrados.append(d)
        elif estado == "Críticos" and filtro_stock_criticos:
            datos_stock_filtrados.append(d)
        elif estado == "Bajos" and filtro_stock_bajos:
            datos_stock_filtrados.append(d)
        elif estado == "Saludables" and filtro_stock_saludables:
            datos_stock_filtrados.append(d)
    
    if datos_stock_filtrados:
        # Crear DataFrame ordenado por estado y nombre
        df_stock = pd.DataFrame(datos_stock_filtrados)
        
        # Ordenar: Críticos primero, luego Agotados, Bajos, Saludables
        estado_orden = {"Críticos": 0, "Agotados": 1, "Bajos": 2, "Saludables": 3}
        df_stock['orden'] = df_stock['Estado'].map(estado_orden)
        df_stock = df_stock.sort_values(['orden', 'Producto']).reset_index(drop=True)
        
        # Obtener productos con stock máximo para calcular porcentaje
        inv_list = api_get("/api/v1/inventario")
        prod_list = api_get("/api/v1/productos")
        
        # CONFIGURACIÓN DE PAGINACIÓN
        productos_por_pagina = 10
        total_productos = len(df_stock)
        total_paginas = (total_productos + productos_por_pagina - 1) // productos_por_pagina
        
        # Inicializar página en session_state si no existe
        if 'pagina_stock' not in st.session_state:
            st.session_state.pagina_stock = 1
        
        # Calcular índices para la página actual
        inicio = (st.session_state.pagina_stock - 1) * productos_por_pagina
        fin = min(inicio + productos_por_pagina, total_productos)
        
        # Crear tabla visual con Streamlit
        st.markdown("""
        <style>
        .stock-table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0 8px;
        }
        .stock-row {
            background: rgba(30, 41, 59, 0.6);
            border-radius: 12px;
            transition: all 0.3s ease;
        }
        .stock-row:hover {
            background: rgba(30, 41, 59, 0.9);
            transform: translateX(5px);
        }
        .stock-cell {
            padding: 16px 20px;
            vertical-align: middle;
        }
        .product-name {
            font-size: 1.1rem;
            font-weight: 600;
            color: #f8fafc;
        }
        .stock-numbers {
            font-size: 1rem;
            color: #94a3b8;
            font-family: 'Cascadia Code', 'Segoe UI', monospace;
        }
        .progress-container {
            width: 200px;
            height: 12px;
            background: rgba(0, 0, 0, 0.4);
            border-radius: 6px;
            overflow: hidden;
            position: relative;
        }
        .progress-fill {
            height: 100%;
            border-radius: 6px;
            transition: width 0.5s ease;
        }
        .status-badge {
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 0.85rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Crear filas de la tabla SOLO para la página actual
        df_pagina = df_stock.iloc[inicio:fin]
        
        for idx, row in df_pagina.iterrows():
            prod = next((p for p in prod_list if p.get('nombre') == row['Producto']), None)
            inv = next((i for i in inv_list if i.get('producto_id') == prod.get('id')), None) if prod else None
            
            stock = row['Stock']
            max_s = prod.get('stock_maximo', 100) if prod else 100
            pct = min((stock / max_s) * 100, 100) if max_s > 0 else 0
            
            # Color según estado
            color_map = {
                "Agotados": {"bg": "#6b7280", "text": "⚫ Agotado"},
                "Críticos": {"bg": "#ef4444", "text": "🔴 Crítico"},
                "Bajos": {"bg": "#f59e0b", "text": "🟡 Bajo"},
                "Saludables": {"bg": "#10b981", "text": "🟢 Saludable"}
            }
            
            estado_info = color_map.get(row['Estado'], color_map["Saludables"])
            
            # Crear fila HTML
            st.markdown(f"""
            <div style="display: flex; align-items: center; background: rgba(30, 41, 59, 0.6); 
                        border-radius: 12px; padding: 16px 20px; margin-bottom: 10px;
                        border-left: 4px solid {estado_info['bg']};
                        transition: all 0.3s ease;">
                <div style="flex: 2; min-width: 150px;">
                    <div style="font-size: 1.1rem; font-weight: 600; color: #f8fafc;">{row['Producto']}</div>
                </div>
                <div style="flex: 1; text-align: center;">
                    <span style="font-size: 1.5rem; font-weight: 600; color: #ffffff; font-family: 'Cascadia Code', 'Segoe UI', monospace;">
                        {stock}
                    </span>
                    <span style="font-size: 1.3rem; font-weight: 700; color: #ffffff;">/</span>
                    <span style="font-size: 1.3rem; font-weight: 700; color: #10b981;">{max_s}</span>
                </div>
                <div style="flex: 1.5; padding: 0 20px;">
                    <div style="width: 100%; height: 12px; background: rgba(0, 0, 0, 0.4); border-radius: 6px; overflow: hidden;">
                        <div style="width: {pct}%; height: 100%; background: {estado_info['bg']}; border-radius: 6px;"></div>
                    </div>
                    <div style="text-align: center; font-size: 1rem; font-weight: 700; color: #ffffff; margin-top: 6px;">{pct:.0f}%</div>
                </div>
                <div style="flex: 1; text-align: right;">
                    <span style="padding: 6px 14px; border-radius: 20px; font-size: 0.85rem; font-weight: 600;
                                background: {estado_info['bg']}30; color: {estado_info['bg']}; border: 1px solid {estado_info['bg']};">
                        {estado_info['text']}
                    </span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # CONTROLES DE PAGINACIÓN
        st.markdown("---")
        col_pag1, col_pag2, col_pag3 = st.columns([1, 2, 1])
        
        with col_pag1:
            if st.session_state.pagina_stock > 1:
                if st.button("⬅️ Anterior", key="btn_prev_stock", use_container_width=True):
                    st.session_state.pagina_stock -= 1
                    st.rerun()
        
        with col_pag2:
            st.markdown(f"""
            <div style="text-align: center; color: #94a3b8; padding: 10px;">
                Página <span style="color: #00f0ff; font-weight: 700; font-size: 1.1rem;">{st.session_state.pagina_stock}</span> 
                de <span style="color: #00f0ff; font-weight: 700; font-size: 1.1rem;">{total_paginas if total_paginas > 0 else 1}</span>
                <br>
                <span style="font-size: 0.85rem;">Mostrando {inicio + 1}-{fin} de {total_productos} productos</span>
            </div>
            """, unsafe_allow_html=True)
        
        with col_pag3:
            if st.session_state.pagina_stock < total_paginas:
                if st.button("Siguiente ➡️", key="btn_next_stock", use_container_width=True):
                    st.session_state.pagina_stock += 1
                    st.rerun()
    else:
        st.info("ℹ️ No hay productos para mostrar con los filtros seleccionados")
    
    with col_chart2:
        st.markdown("""
        <div class="metric-card">
            <h3 style="margin-bottom: 15px;">⚡ Alertas</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Usar los valores calculados del Dashboard (con stock máximo)
        alert_agotados = st.session_state.get('dashboard_agotados', 0)
        alert_criticos = st.session_state.get('dashboard_criticos', 0)
        alert_bajos = st.session_state.get('dashboard_bajos', 0)
        
        if alert_criticos > 0:
            st.markdown(f"""
            <div style="padding: 15px; background: rgba(239, 68, 68, 0.2); border-left: 3px solid #ef4444; border-radius: 0 12px 12px 0; margin-bottom: 10px;">
                <div style="color: #ef4444; font-weight: 600; font-size: 1.1rem;">{alert_criticos} Críticos</div>
                <div style="color: #94a3b8; font-size: 1rem; font-weight: 500; margin-top: 4px;">Reposición urgente</div>
            </div>
            """, unsafe_allow_html=True)
        
        if alert_bajos > 0:
            st.markdown(f"""
            <div style="padding: 15px; background: rgba(245, 158, 11, 0.2); border-left: 3px solid #f59e0b; border-radius: 0 12px 12px 0; margin-bottom: 10px;">
                <div style="color: #f59e0b; font-weight: 600; font-size: 1.1rem;">{alert_bajos} Bajos</div>
                <div style="color: #94a3b8; font-size: 1rem; font-weight: 500; margin-top: 4px;">Revisar pronto</div>
            </div>
            """, unsafe_allow_html=True)
        
        if alert_agotados > 0:
            st.markdown(f"""
            <div style="padding: 15px; background: rgba(107, 114, 128, 0.2); border-left: 3px solid #6b7280; border-radius: 0 12px 12px 0; margin-bottom: 10px;">
                <div style="color: #6b7280; font-weight: 600; font-size: 1.1rem;">{alert_agotados} Agotados</div>
                <div style="color: #94a3b8; font-size: 1rem; font-weight: 500; margin-top: 4px;">Sin existencias</div>
            </div>
            """, unsafe_allow_html=True)
        
        if alert_criticos == 0 and alert_bajos == 0 and alert_agotados == 0:
            st.markdown("""
            <div style="padding: 15px; background: rgba(16, 185, 129, 0.2); border-left: 3px solid #10b981; border-radius: 0 12px 12px 0;">
                <div style="color: #10b981; font-weight: 600; font-size: 1.1rem;">✓ Saludable</div>
                <div style="color: #94a3b8; font-size: 1rem; font-weight: 500; margin-top: 4px;">Inventario en óptimas condiciones</div>
            </div>
            """, unsafe_allow_html=True)
    
    # Recomendaciones
    recomendaciones = resumen.get("recomendaciones", [])
    if recomendaciones:
        st.markdown("---")
        st.markdown("""
        <div class="metric-card">
            <h3>⚠️ Recomendaciones de Reposición</h3>
        </div>
        """, unsafe_allow_html=True)
        
        for rec in recomendaciones[:5]:
            estado = rec.get('estado', 'BAJO')
            color = {"CRÍTICO": "#ef4444", "BAJO": "#f59e0b"}.get(estado, "#10b981")
            st.markdown(f"""
            <div style="padding: 15px; background: var(--bg-card); border-left: 3px solid {color}; border-radius: 0 12px 12px 0; margin-bottom: 10px;">
                <div style="display: flex; justify-content: space-between;">
                    <span style="color: var(--text-primary); font-weight: 600;">{rec.get('producto')}</span>
                    <span style="color: {color}; font-weight: 600;">{estado}</span>
                </div>
                <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-top: 10px; color: var(--text-secondary); font-size: 0.85rem;">
                    <div>Stock: <span style="color: white;">{rec.get('stock_actual', 0)}</span></div>
                    <div>Mín: <span style="color: white;">{rec.get('stock_minimo', 0)}</span></div>
                    <div>Reponer: <span style="color: {color};">{rec.get('cantidad_recomendada', 0)}</span></div>
                </div>
            </div>
            """, unsafe_allow_html=True)


elif menu == "📦 Productos":
    productos = api_get("/api/v1/productos")
    inventarios = api_get("/api/v1/inventario")
    categorias = api_get("/api/v1/categorias")
    proveedores = api_get("/api/v1/proveedores")
    
    # Header con título
    st.markdown("<h2>Gestión de Productos</h2>", unsafe_allow_html=True)
    
    # === BUSCADOR Y FILTROS ===
    st.markdown("""
    <style>
        .filtros-container {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }
        .filtro-search {
            flex: 2;
            min-width: 200px;
        }
        .filtro-select {
            flex: 1;
            min-width: 150px;
        }
        .stTextInput > div > div > input {
            background: rgba(0, 240, 255, 0.1);
            border: 1px solid rgba(0, 240, 255, 0.3);
            color: white;
            border-radius: 10px;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col_busq1, col_busq2, col_busq3 = st.columns([2, 1, 1])
    
    with col_busq1:
        busqueda = st.text_input("🔍 Buscar por nombre o SKU", placeholder="Escribe para buscar...", key="busqueda_prod")
    
    with col_busq2:
        cat_options = ["Todas"] + [c.get("nombre", "Sin categoría") for c in categorias]
        cat_filtro = st.selectbox("Categoría", cat_options, key="filtro_cat")
    
    with col_busq3:
        estado_options = ["Todos", "Agotado", "Crítico", "Bajo", "Saludable"]
        estado_filtro = st.selectbox("Estado stock", estado_options, key="filtro_estado")
    
    # Función para obtener stock y estado
    def get_prod_stock(pid):
        for inv in inventarios:
            if inv.get("producto_id") == pid:
                return inv.get("cantidad", 0)
        return 0
    
    def get_prod_estado(pid, stock):
        stock = stock or 0
        for p in productos:
            if p.get("id") == pid:
                max_s = p.get("stock_maximo", 100) or 100
                if stock <= 0:
                    return "Agotado"
                elif max_s > 0:
                    pct = (stock / max_s) * 100
                    if pct <= 25:
                        return "Crítico"
                    elif pct <= 59:
                        return "Bajo"
                    else:
                        return "Saludable"
                return "Saludable"
        return "Saludable"
    
    # Línea divisoria
    st.markdown("---")
    
    # Inicializar tab activo en session_state si no existe
    if 'producto_tab_activo' not in st.session_state:
        st.session_state.producto_tab_activo = 0  # 0 = Catálogo, 1 = Nuevo, 2 = Edición
    
    # Botones de navegación estilo sidebar
    col_nav1, col_nav2, col_nav3 = st.columns(3)
    
    with col_nav1:
        btn_catalogo = st.button("📋 Catálogo", use_container_width=True, 
                                   type="primary" if st.session_state.producto_tab_activo == 0 else "secondary",
                                   key="btn_nav_catalogo")
        if btn_catalogo:
            st.session_state.producto_tab_activo = 0
            st.rerun()
    
    with col_nav2:
        btn_nuevo = st.button("➕ Nuevo", use_container_width=True,
                              type="primary" if st.session_state.producto_tab_activo == 1 else "secondary",
                              key="btn_nav_nuevo")
        if btn_nuevo:
            st.session_state.producto_tab_activo = 1
            st.rerun()
    
    with col_nav3:
        btn_edicion = st.button("✏️ Edición", use_container_width=True,
                                type="primary" if st.session_state.producto_tab_activo == 2 else "secondary",
                                key="btn_nav_edicion")
        if btn_edicion:
            st.session_state.producto_tab_activo = 2
            st.rerun()
    
    # Botones de exportar debajo, a los lados
    if productos:
        col_exp_izq, col_exp_centro, col_exp_der = st.columns([1, 2, 1])
        with col_exp_izq:
            df_export = pd.DataFrame([{
                "ID": p.get("id"),
                "Nombre": p.get("nombre"),
                "SKU": p.get("sku"),
                "Codigo Barras": p.get("codigo_barras"),
                "Precio Venta": p.get("precio_venta"),
                "Precio Coste": p.get("precio_coste"),
                "Stock": get_prod_stock(p.get("id")),
                "Categoría": next((c.get("nombre") for c in categorias if c.get("id") == p.get("categoria_id")), ""),
                "Proveedor": next((pr.get("nombre") for pr in proveedores if pr.get("id") == p.get("proveedor_id")), ""),
            } for p in productos])
            
            excel_data = to_excel(df_export)
            st.download_button(
                "📊 Excel",
                data=excel_data,
                file_name="productos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_exp_der:
            json_data = df_export.to_json(orient='records', indent=2)
            st.download_button(
                "📄 JSON",
                data=json_data,
                file_name="productos.json",
                mime="application/json",
                use_container_width=True
            )
    
    # Filtrar productos
    productos_filtrados = productos
    
    if busqueda:
        busq_lower = busqueda.lower()
        productos_filtrados = [p for p in productos_filtrados if 
            busq_lower in p.get("nombre", "").lower() or 
            busq_lower in p.get("sku", "").lower()]
    
    if cat_filtro != "Todas":
        cat_id = next((c.get("id") for c in categorias if c.get("nombre") == cat_filtro), None)
        productos_filtrados = [p for p in productos_filtrados if p.get("categoria_id") == cat_id]
    
    if estado_filtro != "Todos":
        productos_filtrados = [p for p in productos_filtrados if get_prod_estado(p.get("id"), get_prod_stock(p.get("id"))) == estado_filtro]
    
    # Mostrar contador
    total_msg = f" de {len(productos)}" if len(productos_filtrados) != len(productos) else ""
    st.markdown(f"""
    <div style="padding: 10px 15px; background: rgba(0, 240, 255, 0.1); border-radius: 10px; margin-bottom: 15px;">
        <span style="color: #00f0ff; font-weight: 600;">{len(productos_filtrados)}</span>
        <span style="color: #94a3b8;"> productos{total_msg}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Mostrar contenido según tab activo
    if st.session_state.producto_tab_activo == 0:
        if productos_filtrados:
            st.markdown("""
            <style>
                .producto-card {
                    background: linear-gradient(135deg, var(--bg-card) 0%, rgba(26, 35, 50, 0.8) 100%);
                    border: 1px solid rgba(0, 240, 255, 0.15);
                    border-radius: 16px;
                    padding: 20px;
                    margin-bottom: 15px;
                    transition: all 0.3s ease;
                }
                .producto-card:hover {
                    border-color: var(--accent-cyan);
                    box-shadow: 0 4px 20px rgba(0, 240, 255, 0.15);
                }
                .producto-nombre {
                    font-size: 1.2rem;
                    font-weight: 600;
                    color: var(--accent-cyan);
                    margin-bottom: 8px;
                }
                .progress-container {
                    background: rgba(0, 0, 0, 0.4);
                    border-radius: 10px;
                    height: 12px;
                    margin-top: 12px;
                    overflow: hidden;
                    position: relative;
                }
                .progress-bar {
                    height: 100%;
                    border-radius: 10px;
                    transition: width 0.5s ease;
                }
                .progress-label {
                    display: flex;
                    justify-content: space-between;
                    font-size: 0.7rem;
                    color: #94a3b8;
                    margin-top: 4px;
                }
</style>
""", unsafe_allow_html=True)
            
            def get_cat_name(cat_id):
                for c in categorias:
                    if c.get("id") == cat_id:
                        return c.get("nombre", "Sin categoría")
                return "Sin categoría"
            
            for prod in productos_filtrados:
                fecha_str = prod.get("fecha_creacion", "")
                if fecha_str:
                    try:
                        fecha_dt = datetime.fromisoformat(fecha_str.replace("Z", "+00:00"))
                        fecha_str = fecha_dt.strftime("%d %b %Y").replace("Apr", "Abr").replace("May", "May").replace("Jun", "Jun")
                    except:
                        fecha_str = "Fecha desconocida"
                
                inv = api_get(f"/api/v1/inventario")
                stock = 0
                for i in inv:
                    if i.get("producto_id") == prod.get("id"):
                        stock = i.get("cantidad", 0)
                        break
                
                min_s = prod.get("stock_minimo", 0)
                max_s = prod.get("stock_maximo", 100) or 100
                
                # Fórmula para ESTADO (color del botón): usar stock MÁXIMO
                if max_s > 0:
                    pct_estado = (stock / max_s) * 100
                else:
                    pct_estado = 100 if stock > 0 else 0
                
                # Fórmula para BARRA (visual): usar stock MÁXIMO
                if max_s > 0:
                    pct_barra = min((stock / max_s) * 100, 100)
                else:
                    pct_barra = 100 if stock > 0 else 0
                
                if stock <= 0:
                    estado_color = "#6b7280"
                    estado_texto = "⚫ Agotado"
                elif pct_estado <= 25:
                    estado_color = "#ef4444"
                    estado_texto = "🔴 Crítico"
                elif pct_estado <= 59:
                    estado_color = "#f59e0b"
                    estado_texto = "🟡 Bajo"
                else:
                    estado_color = "#10b981"
                    estado_texto = "🟢 Saludable"
                
                categoria_nom = get_cat_name(prod.get("categoria_id"))
                
                cat_emoji = {
                    "Lácteos": "🥛", "Lacteos": "🥛", "Bebidas": "🥤", "Frutas": "🍎", "Verduras": "🥬",
                    "Panadería": "🥐", "Panaderia": "🥐", "Carnes": "🥩", "Snacks": "🍿", "General": "📦"
                }
                emoji = cat_emoji.get(categoria_nom, "📦")
                
                img_url = prod.get("imagen_url") or ""
                
                with st.container():
                    col_img, col_info = st.columns([1, 3])
                    
                    with col_img:
                        if img_url:
                            try:
                                st.image(img_url, width=180)
                            except:
                                st.markdown(f"""
                                <div style="width: 180px; height: 180px; background: rgba(0, 240, 255, 0.1); 
                                    border-radius: 12px; display: flex; align-items: center; justify-content: center;
                                    font-size: 3rem;">
                                    {emoji}
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            st.markdown(f"""
                            <div style="width: 180px; height: 180px; background: rgba(0, 240, 255, 0.1); 
                                border-radius: 12px; display: flex; align-items: center; justify-content: center;
                                font-size: 3rem;">
                                {emoji}
                            </div>
                            """, unsafe_allow_html=True)
                        
                        desc = prod.get("descripcion", "")
                        if not desc:
                            descripciones = {
                                "leche": "Leche entera UHT 1.5L clásica de Central Lechera Asturiana, rica en calcio y vitaminas esenciales, perfecta para el desayuno de toda la familia. Formato cómodo con tapón antigoteo, envase reciclable",
                                "huevos": "Docena de huevos frescos de gallinas criadas en libertad, ideales para cualquier receta culinaria. Huevos de categoría A, frescos y nutritivos, perfecta fuente de proteínas",
                                "pan": "Pan artesanal recién horneado cada día con harina de trigo selecta, crujiente por fuera y suave por dentro. Perfecto para sandwiches, tostadas o acompanar cualquier comida",
                                "agua": "Agua mineral natural de fuente pura, sin gases ni impurezas, parfait para hidratarse durante todo el día. Envase plástico reciclable, origen sostenible",
                                "cafe": "Café molido premium de tueste natural con aroma intenso y sabor equilibrado, ideal para empezar el día con energía. Paquete sellado para preservar frescura",
                                "arroz": "Arroz blanco de grano largo tipo basmati, perfecto para arroces tres delicals, paellas y guarniciones. Grano independientes que no se pegan",
                                "aceite": "Aceite de oliva virgen extra de primera presión en frío, color dorado y saborFruit intense, ideal para ensaladas, frituras y aliños. Envase de vidrio oscuro",
                                "azucar": "Azúcar blanco refinado de caña de alta pureza, perfecto para endulzar bebidas calientes frías y recetas de repostería. Textura fina de disolución rápida",
                                "harina": "Harina de trigo multiuso para repostería y panadería, textura suave y esponjosa. Ideal para cakes, pastas,pasteles y panes caseros",
                                "galletas": "Galletas dulces crujientes tipo snack con sabor clásico, ideales para merendar con leche o café. Paquete familiar con cierre hermético para preservar crocancia",
                                "cereal": "Cereal de desayuno integral rico en fibra y vitaminas, fuente de energía natural para la mañana. Crujiente y delicioso con leche",
                            }
                            nombre_prod = prod.get("nombre", "").lower()
                            desc = descripciones.get(nombre_prod, f"Producto de {categoria_nom} de alta calidad, elaborado con los mejores ingredientes para garantizar frescura y sabor excepcional en cada consumo")
                        
                        palabras = desc.split()[:50]
                        desc_corta = " ".join(palabras)
                        st.caption(desc_corta)
                    
                    with col_info:
                        # Botón para seleccionar producto a editar (estilo limpio)
                        col_name, col_arrow = st.columns([4, 1])
                        with col_name:
                            st.markdown(f"""
                            <div style="font-size: 1.3rem; font-weight: 600; color: #00f0ff; 
                                        cursor: pointer; padding: 8px 0;
                                        text-shadow: 0 0 10px rgba(0, 240, 255, 0.3);">
                                {prod.get('nombre', 'Producto')}
                            </div>
                            """, unsafe_allow_html=True)
                        with col_arrow:
                            if st.button("✏️", key=f"select_prod_{prod.get('id')}", help="Editar producto"):
                                st.session_state['editar_producto'] = prod.get('id')
                                st.session_state.producto_tab_activo = 2
                                st.rerun()
                        
                        st.markdown(f"### €{prod.get('precio_venta', 0):.2f}")
                        
                        st.caption(f"📦 SKU: {prod.get('sku', 'N/A')}")
                        st.caption(f"🏷️ {categoria_nom}")
                        
                        st.write(f"📊 **Stock:** {stock} {prod.get('unidad', 'uds')}")
                        
                        st.progress(pct_barra / 100)
                        st.caption(f"{pct_barra:.0f}% del máximo")
                        
                        st.markdown(f"""
                        <div style="padding: 8px 16px; border-radius: 8px; text-align: center; 
                            background: {estado_color}; color: white; font-weight: 600;">
                            {estado_texto}
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown("---")
        else:
            st.info("No hay productos en el catálogo")
    
    elif st.session_state.producto_tab_activo == 1:
        if 'form_version' not in st.session_state:
            st.session_state['form_version'] = 0
        
        categorias = api_get("/api/v1/categorias")
        productos = api_get("/api/v1/productos")
        nombres_existentes = [p.get('nombre', '').lower() for p in productos]
        
        form_version = st.session_state.get('form_version', 0)
        
        st.markdown("""
        <div style="background: linear-gradient(135deg, rgba(0, 240, 255, 0.1), rgba(139, 92, 246, 0.1)); 
                    padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.3); margin-bottom: 20px;">
            <h4 style="margin: 0; color: #00f0ff;">➕ Nuevo Producto</h4>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            sku = st.text_input("Código SKU *", key=f"new_sku_{form_version}", placeholder="Ej: PROD-001")
            nombre = st.text_input("Nombre del producto *", key=f"new_nombre_{form_version}", placeholder="Ej: Leche Entera")
            codigo_barras = st.text_input("Código de barras", key=f"new_codigo_barras_{form_version}", placeholder="Ej: 7622210449283")
            precio = st.number_input("Precio de venta (€) *", min_value=0.0, value=0.0, step=0.01, key=f"new_precio_{form_version}", format="%.2f")
            
        with col2:
            unidad = st.selectbox("Unidad de medida *", ["unidad", "kg", "litro", "paquete", "caja", "botella"], key=f"new_unidad_{form_version}")
            precio_coste = st.number_input("Precio de coste (€)", min_value=0.0, value=0.0, key=f"new_precio_coste_{form_version}")
            tiempo_repo = st.number_input("Días de reposición", min_value=1, value=3, key=f"new_tiempo_{form_version}")
            
            # Proveedor
            prov_list = api_get("/api/v1/proveedores")
            prov_options = {p.get("nombre", "Sin nombre"): p.get("id") for p in prov_list}
            prov_options["Sin proveedor"] = None
            proveedor_nombre = st.selectbox("Proveedor", list(prov_options.keys()), key=f"new_proveedor_{form_version}")
            proveedor_id = prov_options.get(proveedor_nombre)
        
        st.markdown("---")
        st.markdown("<h5 style='color: #00f0ff; margin-bottom: 15px;'>📦 Configuración de Stock</h5>", unsafe_allow_html=True)
        
        # Fila 1: Stock actual | Unidad de ingreso
        col_stock1, col_stock2 = st.columns(2)
        with col_stock1:
            # Stock actual bloqueado en 0
            st.markdown("""
            <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2);">
                <label style="color: #94a3b8; font-size: 0.9rem;">🔒 Stock actual</label>
                <div style="font-size: 1.5rem; font-weight: 700; color: #64748b;">0</div>
                <small style="color: #64748b;">Inicia en 0 (bloqueado)</small>
            </div>
            """, unsafe_allow_html=True)
            stock_actual = 0
            
        with col_stock2:
            # Unidad de ingreso con mismo estilo de card
            st.markdown("""
            <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2);">
                <label style="color: #94a3b8; font-size: 0.9rem;">📥 Unidad de ingreso *</label>
            </div>
            """, unsafe_allow_html=True)
            unidad_ingreso = st.number_input("", min_value=1, value=10, key=f"new_unidad_ingreso_{form_version}",
                                            help="Cuánto entra al reabastecer", label_visibility="collapsed")
        
        # Fila 2: Stock mínimo | Stock máximo
        col_stock3, col_stock4 = st.columns(2)
        with col_stock3:
            # Stock mínimo bloqueado en 0
            st.markdown("""
            <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2); margin-top: 10px;">
                <label style="color: #94a3b8; font-size: 0.9rem;">🔒 Stock mínimo</label>
                <div style="font-size: 1.5rem; font-weight: 700; color: #64748b;">0</div>
                <small style="color: #64748b;">Límite de alerta (bloqueado)</small>
            </div>
            """, unsafe_allow_html=True)
            stock_min = 0
            
        with col_stock4:
            # Stock máximo con mismo estilo de card
            st.markdown("""
            <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2); margin-top: 10px;">
                <label style="color: #94a3b8; font-size: 0.9rem;">📦 Stock máximo *</label>
            </div>
            """, unsafe_allow_html=True)
            stock_max = st.number_input("", min_value=0, value=100, key=f"new_stock_max_{form_version}",
                                       help="Capacidad ideal del almacén", label_visibility="collapsed")
        
        st.markdown("---")
        col3, col4 = st.columns(2)
        with col3:
            cat_options = {c.get("nombre"): c.get("id") for c in categorias}
            categoria_nombre = st.selectbox("Categoría *", list(cat_options.keys()), key=f"new_cat_{form_version}")
            categoria_id = cat_options.get(categoria_nombre)
            
            descripcion = st.text_area("Descripción del producto", key=f"new_descripcion_{form_version}", height=100, 
                                       placeholder="Describe el producto...")
        
        with col4:
            # Foto del producto
            st.markdown("<label style='color: #94a3b8; font-size: 0.9rem; display: block; margin-bottom: 8px;'>📷 Foto del producto</label>", unsafe_allow_html=True)
            
            imagen_subida = st.file_uploader("", type=['png', 'jpg', 'jpeg'], key=f"new_imagen_{form_version}", label_visibility="collapsed")
            
            if imagen_subida is not None:
                st.image(imagen_subida, width=150, caption="Vista previa")
                st.markdown('<div style="color: #10b981; font-size: 0.85rem; margin-top: 8px;">✅ Imagen cargada</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div style="color: #64748b; font-size: 0.85rem; font-style: italic; margin-top: 8px; padding: 10px; background: rgba(30,41,59,0.4); border-radius: 6px; border: 1px dashed rgba(0,240,255,0.3);">📁 Arrastra o haz clic para subir</div>', unsafe_allow_html=True)
        
        # Verificar campos obligatorios
        campos_obligatorios = {
            "Código SKU": sku,
            "Nombre del producto": nombre,
            "Precio de venta": precio >= 0.01,
            "Unidad de medida": unidad,
            "Unidad de ingreso": unidad_ingreso >= 1,
            "Stock máximo": stock_max > 0,
            "Categoría": categoria_id is not None
        }
        
        campos_faltantes = [campo for campo, valor in campos_obligatorios.items() if not valor]
        
        if campos_faltantes:
            st.markdown(f"""
            <div style="background: rgba(239, 68, 68, 0.1); padding: 15px; border-radius: 10px; border: 1px solid rgba(239, 68, 68, 0.3); margin: 20px 0;">
                <span style="color: #ef4444; font-weight: 600;">⚠️ Campos obligatorios faltantes:</span><br>
                <span style="color: #94a3b8;">{', '.join(campos_faltantes)}</span>
            </div>
            """, unsafe_allow_html=True)
            st.button("💾 Crear producto", type="primary", use_container_width=True, key="btn_crear_prod_disabled", disabled=True)
        else:
            st.markdown("<div style='color: #10b981; margin: 20px 0;'>✅ Todos los campos obligatorios completados</div>", unsafe_allow_html=True)
            if st.button("💾 Crear producto", type="primary", use_container_width=True, key="btn_crear_prod"):
                # Validación adicional de precio
                if precio < 0.01:
                    st.error("❌ El precio de venta debe ser mayor a 0.00€")
                elif nombre.lower() in nombres_existentes:
                    st.error("❌ Ya existe un producto con ese nombre")
                else:
                    # Guardar imagen si se subió
                    imagen_url = None
                    if imagen_subida is not None:
                        # Crear directorio si no existe
                        import os
                        os.makedirs("docs/productos", exist_ok=True)
                        
                        # Guardar imagen con nombre único
                        extension = imagen_subida.name.split('.')[-1]
                        nombre_imagen = f"{sku.replace(' ', '_').replace('/', '_')}_{form_version}.{extension}"
                        ruta_imagen = f"docs/productos/{nombre_imagen}"
                        
                        with open(ruta_imagen, "wb") as f:
                            f.write(imagen_subida.getbuffer())
                        
                        imagen_url = ruta_imagen
                        st.success(f"✅ Imagen guardada: {nombre_imagen}")
                    
                    data = {
                        "sku": sku,
                        "nombre": nombre,
                        "codigo_barras": codigo_barras if codigo_barras else None,
                        "precio_venta": precio,
                        "precio_coste": precio_coste if precio_coste > 0 else None,
                        "unidad": unidad,
                        "stock_minimo": stock_min,
                        "stock_maximo": stock_max,
                        "unidad_ingreso": unidad_ingreso,
                        "tiempo_reposicion": tiempo_repo,
                        "categoria_id": categoria_id,
                        "proveedor_id": proveedor_id,
                        "descripcion": descripcion,
                        "imagen_url": imagen_url
                    }
                    result = api_post("/api/v1/productos", data)
                    if result:
                        st.success("✅ Producto añadido al catálogo")
                        prod_id = result.get("id")
                        # Crear inventario inicial con la unidad de ingreso
                        api_post(f"/api/v1/inventario/{prod_id}", {"cantidad": unidad_ingreso, "ubicacion": "Almacén A"})
                        st.session_state['form_version'] = form_version + 1
                        st.rerun()
                    else:
                        st.error("❌ Error al crear el producto. Revisa los datos.")
    
    elif st.session_state.producto_tab_activo == 2:
        # Mostrar mensaje de éxito si existe
        if 'producto_actualizado' in st.session_state and st.session_state['producto_actualizado']:
            st.success("✅ Producto actualizado con éxito")
            del st.session_state['producto_actualizado']
        
        if 'editar_producto' in st.session_state and st.session_state['editar_producto']:
            prod_id_edit = st.session_state['editar_producto']
            producto_edit = next((p for p in productos if p.get('id') == prod_id_edit), None)
            categorias = api_get("/api/v1/categorias")
            proveedores = api_get("/api/v1/proveedores")
            inv_list = api_get("/api/v1/inventario")
            
            if producto_edit:
                # Obtener stock actual del producto
                inv_actual = next((i for i in inv_list if i.get('producto_id') == prod_id_edit), None)
                stock_actual_val = inv_actual.get('cantidad', 0) if inv_actual else 0
                
                # Header con estilo luminoso
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, rgba(0, 240, 255, 0.15), rgba(139, 92, 246, 0.15)); 
                            padding: 25px; border-radius: 15px; border: 1px solid rgba(0, 240, 255, 0.4);
                            box-shadow: 0 0 30px rgba(0, 240, 255, 0.2), inset 0 0 20px rgba(0, 240, 255, 0.05);
                            margin-bottom: 25px; text-align: center;">
                    <h2 style="margin: 0; color: #00f0ff; text-shadow: 0 0 15px rgba(0, 240, 255, 0.6); font-size: 1.8rem;">
                        ✏️ Editando: {producto_edit.get('nombre', '')}
                    </h2>
                    <div style="width: 100%; height: 2px; background: linear-gradient(90deg, transparent, #00f0ff, transparent); 
                                margin-top: 15px; box-shadow: 0 0 10px rgba(0, 240, 255, 0.5);"></div>
                </div>
                """, unsafe_allow_html=True)
                
                col_e1, col_e2 = st.columns(2)
                with col_e1:
                    st.markdown(f"""
                    <div style="margin-bottom: 15px;">
                        <label style="color: #94a3b8; font-size: 0.9rem;">SKU</label>
                        <div style="background: rgba(30, 41, 59, 0.8); padding: 10px 15px; border-radius: 8px; 
                                    border: 1px solid rgba(0, 240, 255, 0.2); color: #64748b; font-weight: 600;">
                            {producto_edit.get('sku', 'N/A')}
                        </div>
                        <small style="color: #64748b;">El SKU no se puede modificar</small>
                    </div>
                    """, unsafe_allow_html=True)
                    edit_nombre = st.text_input("Nombre del producto *", value=producto_edit.get('nombre', ''), key="edit_nombre")
                    edit_codigo_barras = st.text_input("Código de barras", value=producto_edit.get('codigo_barras') or '', key="edit_codigo_barras")
                    edit_precio = st.number_input("Precio de venta (€) *", min_value=0.0, value=float(producto_edit.get('precio_venta', 0)), step=0.01, key="edit_precio", format="%.2f")
                    
                with col_e2:
                    edit_unidad = st.selectbox("Unidad de medida *", ["unidad", "kg", "litro", "paquete", "caja", "botella"], 
                        index=["unidad", "kg", "litro", "paquete", "caja", "botella"].index(producto_edit.get('unidad', 'unidad')) if producto_edit.get('unidad') in ["unidad", "kg", "litro", "paquete", "caja", "botella"] else 0, key="edit_unidad")
                    edit_coste = st.number_input("Precio de coste (€)", value=float(producto_edit.get('precio_coste') or 0), key="edit_coste")
                    edit_tiempo = st.number_input("Días de reposición", value=int(producto_edit.get('tiempo_reposicion', 3)), key="edit_tiempo")
                    
                    # Proveedor
                    prov_options = {p.get("nombre", "Sin nombre"): p.get("id") for p in proveedores}
                    prov_options["Sin proveedor"] = None
                    prov_actual = producto_edit.get('proveedor_id')
                    prov_nombres = list(prov_options.keys())
                    prov_index = next((i for i, (nom, pid) in enumerate(prov_options.items()) if pid == prov_actual), len(prov_nombres) - 1)
                    edit_prov_nombre = st.selectbox("Proveedor", prov_nombres, index=prov_index, key="edit_proveedor")
                    edit_prov_id = prov_options.get(edit_prov_nombre)
                
                st.markdown("---")
                st.markdown("<h5 style='color: #00f0ff; margin-bottom: 15px;'>📦 Configuración de Stock</h5>", unsafe_allow_html=True)
                
                # Fila 1: Stock actual | Unidad de ingreso
                col_stock1, col_stock2 = st.columns(2)
                with col_stock1:
                    # Stock actual editable
                    st.markdown("""
                    <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2);">
                        <label style="color: #94a3b8; font-size: 0.9rem;">📊 Stock actual *</label>
                    </div>
                    """, unsafe_allow_html=True)
                    edit_stock_actual = st.number_input("", min_value=0, value=int(stock_actual_val), key="edit_stock_actual", label_visibility="collapsed")
                    
                with col_stock2:
                    # Unidad de ingreso con mismo estilo de card
                    st.markdown("""
                    <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2);">
                        <label style="color: #94a3b8; font-size: 0.9rem;">📥 Unidad de ingreso</label>
                    </div>
                    """, unsafe_allow_html=True)
                    edit_unidad_ingreso = st.number_input("", min_value=1, value=int(producto_edit.get('unidad_ingreso', 10)), key="edit_unidad_ingreso", label_visibility="collapsed")
                
                # Fila 2: Stock mínimo | Stock máximo
                col_stock3, col_stock4 = st.columns(2)
                with col_stock3:
                    # Stock mínimo bloqueado en 0
                    st.markdown("""
                    <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2); margin-top: 10px;">
                        <label style="color: #94a3b8; font-size: 0.9rem;">🔒 Stock mínimo</label>
                        <div style="font-size: 1.5rem; font-weight: 700; color: #64748b;">0</div>
                        <small style="color: #64748b;">Límite de alerta (bloqueado)</small>
                    </div>
                    """, unsafe_allow_html=True)
                    edit_stock_min = 0
                    
                with col_stock4:
                    # Stock máximo con mismo estilo de card
                    st.markdown("""
                    <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2); margin-top: 10px;">
                        <label style="color: #94a3b8; font-size: 0.9rem;">📦 Stock máximo *</label>
                    </div>
                    """, unsafe_allow_html=True)
                    edit_stock_max = st.number_input("", min_value=0, value=int(producto_edit.get('stock_maximo', 100)), key="edit_stock_max", label_visibility="collapsed")
                
                st.markdown("---")
                col_e3, col_e4 = st.columns(2)
                with col_e3:
                    edit_cat = st.selectbox("Categoría *", [c.get('nombre') for c in categorias], 
                        index=next((i for i, c in enumerate(categorias) if c.get('id') == producto_edit.get('categoria_id')), 0), key="edit_cat")
                    edit_cat_id = next((c.get('id') for c in categorias if c.get('nombre') == edit_cat), 1)
                    
                    edit_descripcion = st.text_area("Descripción del producto", value=producto_edit.get('descripcion') or '', key="edit_descripcion", height=100)
                    
                with col_e4:
                    # Foto del producto en card
                    st.markdown("""
                    <div style="background: rgba(30, 41, 59, 0.6); padding: 15px; border-radius: 10px; border: 1px solid rgba(0, 240, 255, 0.2);">
                        <label style="color: #94a3b8; font-size: 0.9rem;">📷 Foto del producto</label>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Mostrar imagen actual si existe
                    img_actual = producto_edit.get('imagen_url')
                    if img_actual:
                        try:
                            st.image(img_actual, width=150, caption="Imagen actual")
                        except:
                            st.info("No se puede mostrar la imagen actual")
                    
                    edit_imagen = st.file_uploader("Cambiar", type=['png', 'jpg', 'jpeg'], key="edit_imagen", label_visibility="collapsed")
                    if edit_imagen is not None:
                        st.image(edit_imagen, width=150, caption="Nueva imagen")
                        st.success("✅ Nueva imagen cargada")
                
                # Verificar campos obligatorios en edición
                campos_obligatorios_edit = {
                    "Nombre del producto": edit_nombre,
                    "Precio de venta": edit_precio >= 0.01,
                    "Unidad de medida": edit_unidad,
                    "Stock máximo": edit_stock_max > 0,
                    "Categoría": edit_cat_id is not None
                }
                
                campos_faltantes_edit = [campo for campo, valor in campos_obligatorios_edit.items() if not valor]
                
                # Botones arriba
                col_btns = st.columns(2)
                with col_btns[0]:
                    if campos_faltantes_edit:
                        st.button("💾 Guardar cambios", type="primary", key="guardar_prod_disabled", use_container_width=True, disabled=True)
                    else:
                        if st.button("💾 Guardar cambios", type="primary", key="guardar_prod", use_container_width=True):
                            # Validación adicional de precio
                            if edit_precio < 0.01:
                                st.error("❌ El precio de venta debe ser mayor a 0.00€")
                            else:
                                # Guardar nueva imagen si se subió
                                nueva_imagen_url = producto_edit.get('imagen_url')
                                if edit_imagen is not None:
                                    import os
                                    os.makedirs("docs/productos", exist_ok=True)
                                    extension = edit_imagen.name.split('.')[-1]
                                    nombre_imagen = f"{producto_edit.get('sku', 'prod').replace(' ', '_').replace('/', '_')}_edit.{extension}"
                                    ruta_imagen = f"docs/productos/{nombre_imagen}"
                                    with open(ruta_imagen, "wb") as f:
                                        f.write(edit_imagen.getbuffer())
                                    nueva_imagen_url = ruta_imagen
                                
                                data_edit = {
                                "sku": producto_edit.get('sku'),  # Incluir SKU requerido
                                "nombre": edit_nombre,
                                "codigo_barras": edit_codigo_barras if edit_codigo_barras else None,
                                "precio_venta": edit_precio,
                                "precio_coste": edit_coste if edit_coste > 0 else None,
                                "unidad": edit_unidad,
                                "stock_minimo": edit_stock_min,
                                "stock_maximo": edit_stock_max,
                                "unidad_ingreso": edit_unidad_ingreso,
                                "tiempo_reposicion": edit_tiempo,
                                "categoria_id": edit_cat_id,
                                "proveedor_id": edit_prov_id,
                                "descripcion": edit_descripcion,
                                "imagen_url": nueva_imagen_url
                            }
                            result_edit = api_put(f"/api/v1/productos/{prod_id_edit}", data_edit)
                            if result_edit and result_edit.get('id'):
                                # Actualizar stock en inventario
                                api_post(f"/api/v1/inventario/{prod_id_edit}", {"cantidad": edit_stock_actual, "ubicacion": "Almacén A"})
                                # Guardar flag para mostrar mensaje después del rerun
                                st.session_state['producto_actualizado'] = True
                                del st.session_state['editar_producto']
                                st.rerun()
                            else:
                                error_detail = result_edit.get('error', 'Error desconocido')
                                st.error(f"❌ Error al actualizar: {error_detail}")
                with col_btns[1]:
                    if st.button("❌ Cancelar", key="cancelar_edit", use_container_width=True):
                        del st.session_state['editar_producto']
                        st.rerun()
                
                # Mensaje de campos obligatorios abajo a lo largo
                if campos_faltantes_edit:
                    st.markdown(f"""
                    <div style="background: rgba(239, 68, 68, 0.1); padding: 15px; border-radius: 10px; border: 1px solid rgba(239, 68, 68, 0.3); margin-top: 15px;">
                        <span style="color: #ef4444; font-weight: 600;">⚠️ Campos obligatorios faltantes:</span>
                        <span style="color: #94a3b8; margin-left: 10px;">{', '.join(campos_faltantes_edit)}</span>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div style="background: rgba(16, 185, 129, 0.1); padding: 15px; border-radius: 10px; border: 1px solid rgba(16, 185, 129, 0.3); margin-top: 15px;">
                        <span style="color: #10b981; font-weight: 600;">✅ Todos los campos obligatorios completados</span>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info("👆 Selecciona un producto del catálogo para editarlo")


elif menu == "📊 Inventario":
    st.markdown("<h2>Control de Inventario</h2>", unsafe_allow_html=True)
    
    inventarios = api_get("/api/v1/inventario")
    productos = api_get("/api/v1/productos")
    
    # === ACTUALIZAR STOCK ===
    st.markdown("<h3>✏️ Actualizar stock</h3>", unsafe_allow_html=True)
    with st.form("actualizar_stock"):
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            prod_id = st.number_input("ID del producto", min_value=1, key="prod_id")
        with col2:
            nueva_cantidad = st.number_input("Nueva cantidad", min_value=0, key="nueva_cant")
        with col3:
            ubicacion = st.text_input("Ubicación")
        
        if st.form_submit_button("🔄 Actualizar"):
            result = api_post(f"/api/v1/inventario/{prod_id}", {"cantidad": nueva_cantidad, "ubicacion": ubicacion})
            if result:
                st.success("✅ Stock actualizado")
            else:
                st.error("❌ Error")
    
    st.markdown("---")
    
    # === FILTROS ===
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        filtro_estado_inv = st.selectbox("Filtrar por estado", ["Todos", "Agotado", "Crítico", "Bajo", "Saludable"], key="filtro_inv_estado")
    with col_f2:
        ordenar_inv = st.selectbox("Ordenar por", ["Producto (A-Z)", "Stock (mayor)", "Stock (menor)"], key="ordenar_inv")
    
    # Filtrar
    def get_estado_inv(stock, max_s):
        if stock <= 0:
            return "Agotado"
        elif max_s > 0:
            pct = (stock / max_s) * 100
            if pct <= 25:
                return "Crítico"
            elif pct <= 59:
                return "Bajo"
            else:
                return "Saludable"
        return "Saludable"
    
    datos_inv = []
    for inv in inventarios:
        prod = next((p for p in productos if p.get("id") == inv.get("producto_id")), None)
        if prod:
            stock = inv.get("cantidad", 0)
            max_s = prod.get("stock_maximo", 100) or 100
            estado = get_estado_inv(stock, max_s)
            datos_inv.append({
                "producto": prod,
                "stock": stock,
                "min_s": prod.get("stock_minimo", 0),
                "estado": estado,
                "ubicacion": inv.get("ubicacion", "N/A"),
                "max_s": max_s
            })
    
    # Aplicar filtro
    if filtro_estado_inv != "Todos":
        datos_inv = [d for d in datos_inv if d["estado"] == filtro_estado_inv]
    
    # Ordenar
    if ordenar_inv == "Producto (A-Z)":
        datos_inv = sorted(datos_inv, key=lambda x: x["producto"].get("nombre", ""))
    elif ordenar_inv == "Stock (mayor)":
        datos_inv = sorted(datos_inv, key=lambda x: x["stock"], reverse=True)
    elif ordenar_inv == "Stock (menor)":
        datos_inv = sorted(datos_inv, key=lambda x: x["stock"])
    
    # Contador
    st.markdown(f"""
    <div style="padding: 10px 15px; background: rgba(16, 185, 129, 0.1); border-radius: 10px; margin-bottom: 15px;">
        <span style="color: #10b981; font-weight: 600;">{len(datos_inv)}</span>
        <span style="color: #94a3b8;"> productos</span>
    </div>
    """, unsafe_allow_html=True)
    
    # Exportar Excel
    if datos_inv:
        df_inv = pd.DataFrame([{
            "Producto": d["producto"].get("nombre"),
            "SKU": d["producto"].get("sku"),
            "Stock": d["stock"],
            "Stock Mín": d["min_s"],
            "Stock Máx": d["max_s"],
            "Ubicación": d["ubicacion"],
            "Estado": d["estado"]
        } for d in datos_inv])
        
        excel_inv = to_excel(df_inv)
        col_exp, _ = st.columns([1, 3])
        with col_exp:
            st.download_button(
                "📥 Exportar Excel",
                data=excel_inv,
                file_name="inventario.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    st.markdown("""
    <style>
        .inventario-card {
            background: linear-gradient(135deg, var(--bg-card) 0%, rgba(26, 35, 50, 0.8) 100%);
            border: 1px solid rgba(0, 240, 255, 0.15);
            border-radius: 16px;
            padding: 18px;
            margin-bottom: 12px;
            transition: all 0.3s ease;
        }
        .inventario-card.critico { border-left: 4px solid #ef4444; }
        .inventario-card.bajo { border-left: 4px solid #f59e0b; }
        .inventario-card.ok { border-left: 4px solid #10b981; }
        .inventario-nombre {
            font-size: 1.1rem;
            font-weight: 600;
            color: var(--text-primary);
            margin-bottom: 8px;
        }
        .inventario-info {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 8px;
        }
        .inventario-info-item {
            background: rgba(0, 0, 0, 0.3);
            padding: 8px;
            border-radius: 8px;
            text-align: center;
        }
        .inventario-info-label {
            font-size: 0.7rem;
            color: var(--text-secondary);
            text-transform: uppercase;
        }
        .inventario-info-valor {
            font-size: 1.1rem;
            font-weight: 600;
            color: var(--text-primary);
            margin-top: 4px;
        }
        .estado-badge {
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.8rem;
            font-weight: 600;
        }
        .estado-critico { background: rgba(239, 68, 68, 0.2); color: #ef4444; }
        .estado-bajo { background: rgba(245, 158, 11, 0.2); color: #f59e0b; }
        .estado-saludable { background: rgba(16, 185, 129, 0.2); color: #10b981; }
    </style>
    """, unsafe_allow_html=True)
    
    if datos_inv:
        for d in datos_inv:
            prod = d["producto"]
            stock = d["stock"]
            min_s = d["min_s"]
            estado = d["estado"]
            
            cls_estado = {"Crítico": "critico", "Bajo": "bajo", "Saludable": "saludable"}.get(estado, "saludable")
            badge = {"Crítico": "estado-critico", "Bajo": "estado-bajo", "Saludable": "estado-saludable"}.get(estado, "estado-saludable")
            
            prod_id = prod.get("id")
            
            col_card, col_btns = st.columns([4, 1])
            with col_card:
                st.markdown(f"""
                <div class="inventario-card {cls_estado}">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div class="inventario-nombre">{prod.get("nombre", "Producto")}</div>
                        <span class="estado-badge {badge}">{estado}</span>
                    </div>
                    <div class="inventario-info">
                        <div class="inventario-info-item">
                            <div class="inventario-info-label">Stock actual</div>
                            <div class="inventario-info-valor">{stock}</div>
                        </div>
                        <div class="inventario-info-item">
                            <div class="inventario-info-label">Mínimo</div>
                            <div class="inventario-info-valor">{min_s}</div>
                        </div>
                        <div class="inventario-info-item">
                            <div class="inventario-info-label">Máximo</div>
                            <div class="inventario-info-valor">{d.get('max_s', 0)}</div>
                        </div>
                        <div class="inventario-info-item">
                            <div class="inventario-info-label">Ubicación</div>
                            <div class="inventario-info-valor" style="font-size: 0.95rem;">{d.get('ubicacion', 'N/A')}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            with col_btns:
                nuevo_stock = stock
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("−", key=f"menos_{prod_id}", help="Reducir 1"):
                        nuevo_stock = max(0, stock - 1)
                with c2:
                    if st.button("+", key=f"mas_{prod_id}", help="Aumentar 1"):
                        nuevo_stock = stock + 1
                
                if nuevo_stock != stock:
                    result = api_post(f"/api/v1/inventario/{prod_id}", {"cantidad": nuevo_stock, "ubicacion": d.get("ubicacion", "")})
                    if result:
                        st.success("✓ Actualizado")
                    else:
                        st.error("✗ Error")
    

elif menu == "💰 Ventas":
    st.markdown("<h2>Gestión de Ventas</h2>", unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["📋 Historial", "➕ Nueva"])
    
    with tab1:
        ventas = api_get("/api/v1/ventas")
        productos = api_get("/api/v1/productos")
        
        # Exportar Excel
        if ventas:
            df_ventas = pd.DataFrame([{
                "Fecha": v.get("fecha", "")[:10],
                "Producto": get_prod_name(v.get("producto_id")),
                "Cantidad": v.get("cantidad"),
                "Precio Unit.": v.get("precio_unitario"),
                "Total": v.get("cantidad", 0) * v.get("precio_unitario", 0)
            } for v in ventas])
            
            excel_ventas = to_excel(df_ventas)
            col_exp, _ = st.columns([1, 3])
            with col_exp:
                st.download_button(
                    "📥 Exportar Excel",
                    data=excel_ventas,
                    file_name="ventas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        if ventas:
            # Función para obtener nombre del producto
            def get_prod_name(pid):
                for p in productos:
                    if p.get("id") == pid:
                        return p.get("nombre", "Producto")
                return "Producto"
            
            st.markdown("""
            <style>
                .venta-card {
                    background: linear-gradient(135deg, var(--bg-card) 0%, rgba(26, 35, 50, 0.8) 100%);
                    border: 1px solid rgba(0, 240, 255, 0.15);
                    border-radius: 14px;
                    padding: 16px;
                    margin-bottom: 10px;
                }
                .venta-nombre {
                    font-size: 1rem;
                    font-weight: 600;
                    color: var(--accent-cyan);
                    margin-bottom: 6px;
                }
                .venta-detalles {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 8px;
                }
                .venta-info {
                    background: rgba(0, 0, 0, 0.3);
                    padding: 8px;
                    border-radius: 8px;
                    text-align: center;
                }
                .venta-info-label {
                    font-size: 0.7rem;
                    color: var(--text-secondary);
                    text-transform: uppercase;
                }
                .venta-info-valor {
                    font-size: 1rem;
                    font-weight: 600;
                    color: var(--text-primary);
                    margin-top: 4px;
                }
                .total-venta {
                    background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(0, 240, 255, 0.1));
                    border-radius: 10px;
                    padding: 10px;
                    text-align: center;
                    margin-top: 8px;
                }
            </style>
            """, unsafe_allow_html=True)
            
            # Mostrar últimas 20 ventas
            for venta in ventas[:20]:
                # Limpiar fecha
                fecha_str = venta.get("fecha", "")
                if fecha_str:
                    try:
                        fecha_dt = datetime.fromisoformat(fecha_str.replace("Z", "+00:00"))
                        fecha_str = fecha_dt.strftime("%d %b %Y - %H:%M").replace("Apr", "Abr").replace("May", "May")
                    except:
                        fecha_str = "Fecha desconocida"
                
                cantidad = venta.get("cantidad", 0)
                precio = venta.get("precio_unitario", 0)
                total = cantidad * precio
                
                st.markdown(f"""
                <div class="venta-card">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div class="venta-nombre">{get_prod_name(venta.get("producto_id"))}</div>
                        <div style="color: var(--text-secondary); font-size: 0.8rem;">{fecha_str}</div>
                    </div>
                    <div class="venta-detalles">
                        <div class="venta-info">
                            <div class="venta-info-label">Cantidad</div>
                            <div class="venta-info-valor">{cantidad}</div>
                        </div>
                        <div class="venta-info">
                            <div class="venta-info-label">Precio unit.</div>
                            <div class="venta-info-valor">€{precio:.2f}</div>
                        </div>
                        <div class="total-venta">
                            <div class="venta-info-label">Total</div>
                            <div class="venta-info-valor" style="color: var(--accent-green);">€{total:.2f}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No hay ventas registradas")
    
    with tab2:
        productos = api_get("/api/v1/productos")
        
        if productos:
            with st.form("nueva_venta"):
                prod_options = {f"{p.get('nombre')}": p.get("id") for p in productos}
                producto_sel = st.selectbox("Producto", list(prod_options.keys()))
                producto_id = prod_options.get(producto_sel)
                
                col1, col2 = st.columns(2)
                with col1:
                    cantidad = st.number_input("Cantidad", min_value=1, value=1)
                with col2:
                    prod = next((p for p in productos if p.get("id") == producto_id), None)
                    precio = prod.get("precio_venta", 0) if prod else 0
                    st.metric("Precio Unitario", f"€{precio}")
                
                st.markdown(f"""
                <div style="padding: 20px; background: linear-gradient(135deg, rgba(0, 240, 255, 0.1), rgba(139, 92, 246, 0.1)); border-radius: 12px; text-align: center;">
                    <div style="font-size: 2rem; color: var(--accent-cyan);">Total: €{cantidad * precio:.2f}</div>
                </div>
                """, unsafe_allow_html=True)
                
                if st.form_submit_button("💰 Registrar Venta"):
                    result = api_post("/api/v1/ventas", {"producto_id": producto_id, "cantidad": cantidad, "precio_unitario": precio})
                    if result:
                        st.success("✅ Venta registrada correctamente")
                    else:
                        st.error("❌ Error")


elif menu == "🔮 Predicciones":
    st.markdown("<h2>Predicciones de Demanda</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color: var(--text-secondary);'>Basado en el historial de ventas y algoritmos de ML</p>", unsafe_allow_html=True)
    
    predicciones = api_get("/api/v1/prediccion/todos")
    
    if predicciones:
        datos_grafico = []
        for pred in predicciones:
            datos_grafico.append({
                "Producto": pred.get("producto"),
                "Días Restantes": pred.get("dias_hasta_agotarse"),
                "Estado": pred.get("estado"),
                "Tendencia": pred.get("tendencia")
            })
        
        df = pd.DataFrame(datos_grafico)
        df = df.sort_values("Días Restantes")
        
        colors = {"CRÍTICO": "#ef4444", "BAJO": "#f59e0b", "MODERADO": "#3b82f6", "ADECUADO": "#10b981"}
        
        fig = px.bar(
            df.head(10),
            y="Producto",
            x="Días Restantes",
            color="Estado",
            color_discrete_map=colors,
            orientation="h",
            title="Días hasta agotarse (Top 10 productos)"
        )
        fig.update_layout(height=400, paper_bgcolor="#0a0e17", plot_bgcolor="#0a0e17", font=dict(color="white"))
        st.plotly_chart(fig, width='stretch')
    else:
        st.info("Registra ventas para ver predicciones")


elif menu == "📸 Visión AI":
    st.markdown("<h2>Detección con YOLOv8</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color: var(--text-secondary);'>Sube una imagen de estantería para detectar productos y actualizar inventario</p>", unsafe_allow_html=True)
    
    archivo = st.file_uploader("📁 Seleccionar imagen", type=["jpg", "png", "jpeg"])
    
    col1, col2 = st.columns([1, 3])
    with col1:
        confianza = st.slider("Nivel de confianza", 0.05, 0.5, 0.15)
    with col2:
        actualizar = st.checkbox("Actualizar inventario automáticamente", value=True)
    
    if archivo:
        col_img1, col_img2 = st.columns(2)
        
        with col_img1:
            st.markdown("""
            <div class="metric-card">
                <h4>📷 Imagen Analizada</h4>
            </div>
            """, unsafe_allow_html=True)
            st.image(archivo, caption="Imagen seleccionada", width='stretch')
        
        with col_img2:
            st.markdown("""
            <div class="metric-card">
                <h4>🔍 Resultados</h4>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("⚡ Detectar y Actualizar", type="primary"):
                with st.spinner("🤖 Analizando con YOLOv8..."):
                    files = {"archivo": archivo}
                    data = {"confianza_min": confianza, "actualizar_stock": actualizar}
                    
                    try:
                        response = requests.post(
                            f"{API_URL}/api/v1/vision/analizar-y-actualizar",
                            files=files,
                            data=data,
                            timeout=60
                        )
                        
                        if response.status_code == 200:
                            resultado = response.json()
                            deteccion = resultado.get("deteccion", {})
                            total = deteccion.get("total_objetos", 0)
                            
                            st.markdown(f"""
                            <div style="padding: 20px; background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(0, 240, 255, 0.1)); border-radius: 12px; text-align: center; margin-bottom: 20px;">
                                <div style="font-size: 2rem;">🎯</div>
                                <div style="font-size: 1.5rem; color: var(--accent-cyan);">Detectados {total} objetos</div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            objetos = deteccion.get("objetos", [])
                            if objetos:
                                for obj in objetos:
                                    conf = obj['confianza'] * 100
                                    st.markdown(f"""
                                    <div style="padding: 10px; background: var(--bg-card); border-radius: 8px; margin-bottom: 5px;">
                                        <span style="color: var(--accent-cyan);">{obj['nombre']}</span>
                                        <span style="float: right; color: var(--accent-purple);">{conf:.1f}%</span>
                                    </div>
                                    """, unsafe_allow_html=True)
                            
                            mapeo = resultado.get("mapeo", {})
                            mapeados = mapeo.get("mapeados", [])
                            
                            if mapeados:
                                st.markdown("---")
                                st.markdown("""
                                <div style="padding: 15px; background: rgba(16, 185, 129, 0.2); border-radius: 12px;">
                                    <h4 style="color: var(--accent-green); margin: 0;">✓ Inventario Actualizado</h4>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                for m in mapeados:
                                    st.markdown(f"""
                                    <div style="padding: 8px; margin-left: 20px;">
                                        • {m['producto_nombre']}: <span style="color: var(--accent-cyan);">{m['cantidad_detectada']} uds</span>
                                    </div>
                                    """, unsafe_allow_html=True)
                        else:
                            st.error(f"Error: {response.status_code}")
                    except Exception as e:
                        st.error(f"Error: {str(e)}")


elif menu == "🔍 Barcode":
    st.markdown("<h2>🔍 Búsqueda por Código de Barras</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color: var(--text-secondary);'>Busca productos en la base de datos OpenFoodFacts</p>", unsafe_allow_html=True)
    
    if 'barcode_value' not in st.session_state:
        st.session_state.barcode_value = ""
    
    col_busq1, col_busq2 = st.columns([2, 1])
    with col_busq1:
        barcode = st.text_input("📊 Ingresa el código de barras", placeholder="Ej: 7622210449283", key="barcode_input", value=st.session_state.barcode_value)
    with col_busq2:
        st.markdown("<br>", unsafe_allow_html=True)
        buscar_btn = st.button("🔍 Buscar", use_container_width=True)
    
    if barcode:
        st.session_state.barcode_value = barcode
    
    if buscar_btn and barcode:
        with st.spinner("Buscando en OpenFoodFacts..."):
            try:
                headers = {
                    'User-Agent': 'MarkeTTalento/1.0 (Streamlit Inventory App)',
                    'Accept': 'application/json'
                }
                response = requests.get(
                    f"https://world.openfoodfacts.org/api/v0/product/{barcode}.json",
                    timeout=15,
                    headers=headers
                )
                
                if response.status_code == 200:
                    data = response.json()
                    st.write(f"Debug: Status {response.status_code}, Data: {data.get('status')}")  # Debug
                    if data.get("status") == 1:
                        product = data.get("product", {})
                        
                        st.session_state.barcode_product = product
                        
                        st.markdown("---")
                        
                        col_prod1, col_prod2 = st.columns([1, 2])
                        
                        with col_prod1:
                            if product.get("image_url"):
                                st.image(product.get("image_url"), width=200)
                            else:
                                st.markdown("""
                                <div style="width: 200px; height: 200px; background: var(--bg-card); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                                    <span style="font-size: 3rem;">📦</span>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        with col_prod2:
                            nombre_default = product.get("product_name", "Producto")
                            st.markdown(f"**Nombre:** {nombre_default}")
                            st.markdown(f"**Marca:** {product.get('brands', 'N/A')}")
                            
                            col_cat = st.selectbox("Categoría", ["General", "Bebidas", "Lácteos", "Frutas", "Verduras", "Panadería", "Carnes", "Snacks"], index=0, key="cat_select")
                            col_pre1, col_pre2 = st.columns(2)
                            with col_pre1:
                                precio_v = st.number_input("Precio venta", min_value=0.01, value=10.0, key="pv_input")
                            with col_pre2:
                                precio_c = st.number_input("Precio coste", min_value=0.0, value=5.0, key="pc_input")
                            
                            sku = f"OFF-{barcode[:6]}"
                            
                            if st.button("➕ AGREGAR A PRODUCTOS", type="primary", use_container_width=True, key="btn_agregar_final"):
                                st.session_state.agregar_click = True
                                st.session_state.producto_data = {
                                    "product": product,
                                    "barcode": barcode,
                                    "sku": sku,
                                    "nombre": nombre_default,
                                    "categoria": col_cat,
                                    "precio_v": float(precio_v),
                                    "precio_c": float(precio_c)
                                }
                            
                            if st.session_state.get("agregar_click"):
                                datos = st.session_state.producto_data
                                prod = datos["product"]
                                cat_map = {"General": 1, "Bebidas": 2, "Lácteos": 3, "Frutas": 4, "Verduras": 5, "Panadería": 6, "Carnes": 7, "Snacks": 8}
                                cat_id = cat_map.get(datos["categoria"], 1)
                                
                                nuevo = {
                                    "sku": datos["sku"],
                                    "codigo_barras": datos["barcode"],
                                    "nombre": datos["nombre"][:200],
                                    "descripcion": f"OpenFoodFacts. Marca: {prod.get('brands', 'N/A')}",
                                    "precio_venta": datos["precio_v"],
                                    "precio_coste": datos["precio_c"] if datos["precio_c"] > 0 else None,
                                    "unidad": "unidad",
                                    "stock_minimo": 5,
                                    "stock_maximo": 50,
                                    "tiempo_reposicion": 3,
                                    "categoria_id": cat_id,
                                    "proveedor_id": None,
                                    "imagen_url": prod.get("image_url")
                                }
                                
                                resp = requests.post(f"{API_URL}/api/v1/productos", json=nuevo, timeout=15)
                                if resp.status_code in [200, 201]:
                                    st.success(f"✓ '{datos['nombre']}' AGREGADO!")
                                    st.balloons()
                                    st.session_state.agregar_click = False
                                else:
                                    st.error(f"Error: {resp.status_code}")
                    else:
                        st.error("❌ Producto no encontrado en la base de datos")
                else:
                    st.error(f"Error al conectar con OpenFoodFacts (código: {response.status_code})")
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    if not barcode or not buscar_btn:
        st.markdown("""
        <div style="text-align: center; padding: 40px; color: var(--text-secondary);">
            <p style="font-size: 4rem; margin: 0;">📊</p>
            <p>Ingresa un código de barras para buscar información del producto</p>
            <p style="font-size: 0.85rem; color: var(--accent-purple);">Base de datos: OpenFoodFacts</p>
        </div>
        """, unsafe_allow_html=True)


st.markdown("""
<div style="text-align: center; padding: 20px 0; color: var(--text-secondary); border-top: 1px solid rgba(0, 240, 255, 0.1); margin-top: 30px;">
    <p style="margin: 0;">MarkeTTalento v1.0 - Sistema de Inventario Inteligente</p>
    <p style="font-size: 0.8rem; margin-top: 5px;">Powered by YOLOv8 • FastAPI • Streamlit</p>
</div>
""", unsafe_allow_html=True)
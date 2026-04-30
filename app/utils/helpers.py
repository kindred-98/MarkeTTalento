"""
Funciones auxiliares y helpers
"""
import base64
from io import BytesIO
from typing import Any, Dict, List, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def to_excel(df: Optional[pd.DataFrame]) -> bytes:
    """Convierte un DataFrame a formato Excel con estilos."""
    if df is None or df.empty:
        return b""
    
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    try:
        ws.title = "Datos"
    except Exception:
        pass
    
    # Estilos
    header_fill = PatternFill(start_color="00f0ff", end_color="00f0ff", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir datos
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Estilo header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Estilo celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
    
    # Auto-ajustar columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    wb.save(output)
    return output.getvalue()


def format_currency(value: float) -> str:
    """Formatea un valor como moneda (€)."""
    return f"€{value:.2f}"


def format_date(date_str: str) -> str:
    """Formatea una fecha ISO a formato legible."""
    from datetime import datetime
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return dt.strftime("%d %b %Y").replace("Apr", "Abr").replace("May", "May")
    except Exception:
        return "Fecha desconocida"


def truncate_text(text: str, max_words: int = 50) -> str:
    """Trunca un texto a un número máximo de palabras."""
    if not text:
        return ""
    words = text.split()[:max_words]
    return " ".join(words)


def calcular_porcentaje(actual: int, maximo: int) -> float:
    """Calcula el porcentaje de stock."""
    if maximo <= 0:
        return 100 if actual > 0 else 0
    return min((actual / maximo) * 100, 100)


def get_prod_name(producto_id: int, productos: List[Dict[str, Any]]) -> str:
    """Obtiene el nombre de un producto por su ID."""
    for p in productos:
        if p.get("id") == producto_id:
            return p.get("nombre", "Producto")
    return "Producto"
